from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render
import csv
from openpyxl import Workbook
from inventory.models import Material, Tool
from activities.models import Project, Activity

@login_required
def reports_home(request):
    return render(request, 'reports/reports_home.html')

@login_required
def inventory_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="inventario.csv"'
    writer = csv.writer(response)
    writer.writerow(['Material','Unidad','Stock','Stock Minimo'])
    for m in Material.objects.all():
        writer.writerow([m.name, m.unit, m.stock, m.min_stock])
    writer.writerow([])
    writer.writerow(['Herramienta','Código','Estado'])
    for t in Tool.objects.all():
        writer.writerow([t.name, t.code, t.status])
    return response

@login_required
def activities_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="actividades.csv"'
    writer = csv.writer(response)
    writer.writerow(['Proyecto','Actividad','Progreso','Estado'])
    for a in Activity.objects.select_related('project').all():
        writer.writerow([a.project.name, a.name, a.progress_percent, a.status])
    return response

@login_required
def inventory_excel(request):
    from openpyxl.styles import Border, Side, Alignment, PatternFill, Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="FFD3D3D3", end_color="FFD3D3D3", fill_type="solid")
    header_font = Font(bold=True)

    # Materiales
    ws['A1'] = 'Material'
    ws['B1'] = 'Unidad'
    ws['C1'] = 'Stock'
    ws['D1'] = 'Stock Mínimo'

    for col in ['A1', 'B1', 'C1', 'D1']:
        ws[col].border = thin_border
        ws[col].fill = header_fill
        ws[col].font = header_font
        ws[col].alignment = Alignment(horizontal='center')

    row = 2
    for m in Material.objects.all():
        ws[f'A{row}'] = m.name
        ws[f'B{row}'] = m.get_unit_display()
        ws[f'C{row}'] = m.stock
        ws[f'D{row}'] = m.min_stock
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].border = thin_border
        row += 1

    # Espacio
    row += 1

    # Herramientas
    ws[f'A{row}'] = 'Herramienta'
    ws[f'B{row}'] = 'Código'
    ws[f'C{row}'] = 'Estado'

    for col in ['A', 'B', 'C']:
        cell = ws[f'{col}{row}']
        cell.border = thin_border
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    row += 1
    for t in Tool.objects.all():
        ws[f'A{row}'] = t.name
        ws[f'B{row}'] = t.code
        ws[f'C{row}'] = t.get_status_display()
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].border = thin_border
        row += 1

    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="inventario.xlsx"'
    wb.save(response)
    return response

@login_required
def activities_excel(request):
    from openpyxl.styles import Border, Side, Alignment, PatternFill, Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Actividades"

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="FFD3D3D3", end_color="FFD3D3D3", fill_type="solid")
    header_font = Font(bold=True)

    ws['A1'] = 'Proyecto'
    ws['B1'] = 'Actividad'
    ws['C1'] = 'Progreso (%)'
    ws['D1'] = 'Estado'

    for col in ['A1', 'B1', 'C1', 'D1']:
        ws[col].border = thin_border
        ws[col].fill = header_fill
        ws[col].font = header_font
        ws[col].alignment = Alignment(horizontal='center')

    row = 2
    for a in Activity.objects.select_related('project').all():
        ws[f'A{row}'] = a.project.name
        ws[f'B{row}'] = a.name
        ws[f'C{row}'] = a.progress_percent
        ws[f'D{row}'] = a.get_status_display()
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].border = thin_border
        row += 1

    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="actividades.xlsx"'
    wb.save(response)
    return response
